from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from dotenv import load_dotenv
from flask.cli import with_appcontext
import click

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Session Configuration - Sessions expire on server restart
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)

# Email Configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', True)
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'your-email@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', 'your-app-password')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@attendance.com')

# Generate unique server instance ID on startup
import uuid
SERVER_INSTANCE_ID = str(uuid.uuid4())

db = SQLAlchemy(app)
mail = Mail(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ===================== Models =====================
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), default='employee')  # 'admin' or 'employee'
    department = db.Column(db.String(120))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    attendance_records = db.relationship('Attendance', backref='user', lazy=True, cascade='all, delete-orphan')
    rotas = db.relationship('Rota', backref='user', lazy=True, cascade='all, delete-orphan')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username}>'


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    check_in = db.Column(db.DateTime)
    check_out = db.Column(db.DateTime)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='present')  # 'present', 'absent', 'leave'
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Attendance {self.user_id} - {self.date}>'


class Rota(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    day_of_week = db.Column(db.String(20), nullable=False)  # 'Monday', 'Tuesday', etc.
    shift_start = db.Column(db.Time, nullable=False)
    shift_end = db.Column(db.Time, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Rota {self.user_id} - {self.day_of_week}>'


# ===================== Login Manager =====================
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ===================== Request Handlers =====================
@app.before_request
def check_server_instance():
    """Check if server was restarted - logout users if instance changed"""
    if current_user.is_authenticated and request.endpoint != 'logout':
        # Check if session has the current server instance ID
        if 'server_instance_id' not in session or session.get('server_instance_id') != SERVER_INSTANCE_ID:
            logout_user()
            session.clear()
            return redirect(url_for('login'))


# ===================== Routes =====================
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password) and user.is_active:
            login_user(user)
            # Store server instance ID in session
            session['server_instance_id'] = SERVER_INSTANCE_ID
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('employee_dashboard'))
        else:
            return render_template('login.html', error='Invalid username or password')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        password = request.form.get('password')
        
        current_user.full_name = full_name
        
        # Only update email if it's changed and not already in use
        if email != current_user.email:
            if User.query.filter_by(email=email).first():
                return render_template('profile.html', error='Email already in use', user=current_user)
            current_user.email = email
        
        # Only update password if provided
        password_changed = False
        if password:
            current_user.set_password(password)
            password_changed = True
        
        db.session.commit()
        
        # Send password change notification
        if password_changed:
            send_password_change_email(current_user)
        
        return render_template('profile.html', success='Profile updated successfully', user=current_user)
    
    return render_template('profile.html', user=current_user)


@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    if current_user.role != 'employee':
        return redirect(url_for('index'))
    
    today = datetime.utcnow().date()
    today_attendance = Attendance.query.filter_by(
        user_id=current_user.id,
        date=today
    ).first()
    
    # Get today's rota
    current_day = datetime.utcnow().strftime('%A')
    today_rota = Rota.query.filter_by(
        user_id=current_user.id,
        day_of_week=current_day,
        is_active=True
    ).first()
    
    # Get all rotas for the week
    all_rotas = Rota.query.filter_by(
        user_id=current_user.id,
        is_active=True
    ).order_by(
        db.case(
            (Rota.day_of_week == 'Monday', 1),
            (Rota.day_of_week == 'Tuesday', 2),
            (Rota.day_of_week == 'Wednesday', 3),
            (Rota.day_of_week == 'Thursday', 4),
            (Rota.day_of_week == 'Friday', 5),
            (Rota.day_of_week == 'Saturday', 6),
            (Rota.day_of_week == 'Sunday', 7),
        )
    ).all()
    
    return render_template('employee_dashboard.html', 
                         attendance=today_attendance,
                         today_rota=today_rota,
                         all_rotas=all_rotas)


@app.route('/employee/check-in', methods=['POST'])
@login_required
def check_in():
    if current_user.role != 'employee':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    now = datetime.utcnow()
    today = now.date()
    current_time = now.time()
    current_day = now.strftime('%A')
    
    # Check if employee has a rota for today
    rota = Rota.query.filter_by(
        user_id=current_user.id,
        day_of_week=current_day,
        is_active=True
    ).first()
    
    if not rota:
        return jsonify({'success': False, 'message': f'No schedule assigned for {current_day}. Please contact admin.'})
    
    # Check if current time is within the allowed shift time (30 minutes before shift start)
    shift_start = rota.shift_start
    early_check_in = (datetime.combine(today, shift_start) - timedelta(minutes=30)).time()
    shift_end = rota.shift_end
    
    if current_time < early_check_in:
        return jsonify({'success': False, 'message': f'Too early to check in. Your shift starts at {shift_start.strftime("%H:%M")}. You can check in 30 minutes before.'})
    
    if current_time > shift_end:
        return jsonify({'success': False, 'message': f'Your shift ended at {shift_end.strftime("%H:%M")}. Cannot check in after shift end.'})
    
    existing = Attendance.query.filter_by(
        user_id=current_user.id,
        date=today
    ).first()

    if existing:
        # Check if shift times have changed - if both check_in and check_out exist but shift times don't match
        if existing.check_in and existing.check_out:
            # Employee has completed a previous shift, but there's a new shift assigned
            # Allow them to check in for the new shift by resetting the attendance
            existing.check_in = now
            existing.check_out = None
            existing.status = 'present'
        elif existing.check_in:
            return jsonify({'success': False, 'message': 'Already checked in today'})
        else:
            existing.check_in = now
            existing.status = 'present'
    else:
        attendance = Attendance(
            user_id=current_user.id,
            date=today,
            check_in=now,
            status='present'
        )
        db.session.add(attendance)

    db.session.commit()
    return jsonify({'success': True, 'message': 'Check-in successful', 'time': now.strftime('%H:%M:%S')})


@app.route('/employee/check-out', methods=['POST'])
@login_required
def check_out():
    if current_user.role != 'employee':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    today = datetime.utcnow().date()
    attendance = Attendance.query.filter_by(
        user_id=current_user.id,
        date=today
    ).first()

    if not attendance:
        return jsonify({'success': False, 'message': 'No check-in record found'})

    if attendance.check_out:
        return jsonify({'success': False, 'message': 'Already checked out today'})

    attendance.check_out = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'message': 'Check-out successful', 'time': datetime.utcnow().strftime('%H:%M:%S')})


@app.route('/employee/my-records')
@login_required
def my_records():
    if current_user.role != 'employee':
        return redirect(url_for('index'))
    
    page = request.args.get('page', 1, type=int)
    records = Attendance.query.filter_by(user_id=current_user.id).order_by(
        Attendance.date.desc()
    ).paginate(page=page, per_page=10)
    
    return render_template('my_records.html', records=records)


@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    total_employees = User.query.filter_by(role='employee').count()
    today = datetime.utcnow().date()
    present_today = db.session.query(Attendance).filter(
        Attendance.date == today,
        Attendance.status == 'present'
    ).count()
    
    return render_template('admin_dashboard.html', 
                         total_employees=total_employees,
                         present_today=present_today)


@app.route('/admin/employees')
@login_required
def manage_employees():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    # Filters
    page = request.args.get('page', 1, type=int)
    username_q = request.args.get('username', '', type=str)
    role_filter = request.args.get('role', 'employee', type=str)
    status_filter = request.args.get('status', '', type=str)  # '', 'active', 'inactive'

    query = User.query

    # Role filter (admin or employee)
    if role_filter in ('admin', 'employee'):
        query = query.filter(User.role == role_filter)

    # Username search
    if username_q:
        like = f"%{username_q}%"
        query = query.filter((User.username.ilike(like)) | (User.full_name.ilike(like)) | (User.email.ilike(like)))

    # Status filter
    if status_filter == 'active':
        query = query.filter(User.is_active.is_(True))
    elif status_filter == 'inactive':
        query = query.filter(User.is_active.is_(False))

    query = query.order_by(User.full_name.asc())
    employees = query.paginate(page=page, per_page=10)

    return render_template('manage_employees.html', 
                           employees=employees,
                           username_q=username_q,
                           role_filter=role_filter,
                           status_filter=status_filter)

@app.route('/admin/employees/delete-bulk', methods=['POST'])
@login_required
def delete_employees_bulk():
    if current_user.role != 'admin':
        return redirect(url_for('index'))

    ids = request.form.getlist('ids')
    deleted = 0
    for id_str in ids:
        try:
            uid = int(id_str)
        except ValueError:
            continue

        user = User.query.get(uid)
        if not user:
            continue
        # Skip admins and self
        if user.role == 'admin' or user.id == current_user.id:
            continue
        db.session.delete(user)
        deleted += 1

    if deleted:
        db.session.commit()

    return redirect(url_for('manage_employees'))


@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    user = User.query.get(employee_id)
    if not user:
        return jsonify({'success': False, 'message': 'Employee not found'}), 404
    
    # Prevent deleting admins or self
    if user.role == 'admin':
        return jsonify({'success': False, 'message': 'Cannot delete admin users'}), 403
    
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'Cannot delete yourself'}), 403
    
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Employee deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/employee/<int:employee_id>')
@login_required
def view_employee(employee_id):
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    employee = User.query.get_or_404(employee_id)
    if employee.role != 'employee':
        return redirect(url_for('manage_employees'))
    
    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)
    
    records = Attendance.query.filter_by(user_id=employee_id).filter(
        db.extract('month', Attendance.date) == month,
        db.extract('year', Attendance.date) == year
    ).order_by(Attendance.date.desc()).all()
    
    return render_template('view_employee.html', employee=employee, records=records, month=month, year=year)
# ===================== Helper Functions =====================
def send_welcome_email(user, password):
    """Send welcome email to newly created employee"""
    try:
        subject = f"Welcome to D Attendance System - Your Account Details"
        html_body = f"""
        <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #667EEA 0%, #764BA2 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                    .content {{ background: #f9f9f9; padding: 20px; border-radius: 0 0 8px 8px; border: 1px solid #ddd; }}
                    .info-box {{ background: white; padding: 15px; margin: 15px 0; border-left: 4px solid #667EEA; border-radius: 4px; }}
                    .credentials {{ background: #e8f4f8; padding: 15px; border-radius: 4px; margin: 15px 0; }}
                    .credentials p {{ margin: 8px 0; font-family: monospace; }}
                    .footer {{ text-align: center; margin-top: 20px; font-size: 12px; color: #999; }}
                    .button {{ display: inline-block; background: #667EEA; color: white; padding: 10px 20px; border-radius: 4px; text-decoration: none; margin: 15px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Welcome to D Attendance System! 👋</h1>
                    </div>
                    <div class="content">
                        <p>Hello <strong>{user.full_name}</strong>,</p>
                        
                        <p>Your employee account has been successfully created in the D Attendance System. Your administrator has provided you with the credentials below to get started.</p>
                        
                        <div class="info-box">
                            <h3>🔐 Your Login Credentials</h3>
                            <div class="credentials">
                                <p><strong>Username:</strong> <code>{user.username}</code></p>
                                <p><strong>Password:</strong> <code>{password}</code></p>
                                <p><strong>Email:</strong> <code>{user.email}</code></p>
                            </div>
                        </div>
                        
                        <div class="info-box">
                            <h3>ℹ️ Important Information</h3>
                            <ul>
                                <li><strong>Change Your Password:</strong> Please change your password on first login by going to your Profile settings.</li>
                                <li><strong>Keep Credentials Safe:</strong> Never share your login credentials with anyone.</li>
                                <li><strong>System Features:</strong> You can check your attendance, view rotas, and submit records using this system.</li>
                            </ul>
                        </div>
                        
                        <p style="text-align: center; margin-top: 30px;">
                            <a href="#" class="button">Login to System</a>
                        </p>
                        
                        <p>If you have any questions or issues logging in, please contact your administrator.</p>
                        
                        <p>Best regards,<br><strong>D Attendance System</strong></p>
                    </div>
                    <div class="footer">
                        <p>This is an automated message. Please do not reply to this email.</p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = Message(subject=subject, recipients=[user.email], html=html_body)
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send email to {user.email}: {str(e)}")
        return False


def send_password_change_email(user):
    """Send notification email when password is changed"""
    try:
        subject = f"Password Changed - D Attendance System"
        html_body = f"""
        <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                    .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                    .header {{ background: linear-gradient(135deg, #667EEA 0%, #764BA2 100%); color: white; padding: 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                    .content {{ background: #f9f9f9; padding: 20px; border-radius: 0 0 8px 8px; border: 1px solid #ddd; }}
                    .info-box {{ background: white; padding: 15px; margin: 15px 0; border-left: 4px solid #667EEA; border-radius: 4px; }}
                    .alert-box {{ background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; border-radius: 4px; margin: 15px 0; }}
                    .footer {{ text-align: center; margin-top: 20px; font-size: 12px; color: #999; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Password Changed Successfully 🔐</h1>
                    </div>
                    <div class="content">
                        <p>Hello <strong>{user.full_name}</strong>,</p>
                        
                        <p>This email confirms that your password for the D Attendance System has been successfully changed.</p>
                        
                        <div class="info-box">
                            <h3>📋 Change Details</h3>
                            <p><strong>Account:</strong> {user.username}</p>
                            <p><strong>Email:</strong> {user.email}</p>
                            <p><strong>Date & Time:</strong> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
                        </div>
                        
                        <div class="alert-box">
                            <h3>⚠️ Didn't make this change?</h3>
                            <p>If you did not request this password change, please contact your system administrator immediately to secure your account.</p>
                        </div>
                        
                        <div class="info-box">
                            <h3>💡 Security Tips</h3>
                            <ul>
                                <li>Use a strong, unique password</li>
                                <li>Never share your password with anyone</li>
                                <li>Change your password regularly</li>
                                <li>Log out from shared computers</li>
                            </ul>
                        </div>
                        
                        <p>Best regards,<br><strong>D Attendance System</strong></p>
                    </div>
                    <div class="footer">
                        <p>This is an automated security notification. Please do not reply to this email.</p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        msg = Message(subject=subject, recipients=[user.email], html=html_body)
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send email to {user.email}: {str(e)}")
        return False


# ===================== Routes =====================

@app.route('/admin/add-employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    employee_id = request.args.get('id', type=int)
    employee = None
    
    # If editing, fetch the employee
    if employee_id:
        employee = User.query.get_or_404(employee_id)
        if employee.role == 'admin':
            return redirect(url_for('manage_employees'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        full_name = request.form.get('full_name')
        password = request.form.get('password')
        department = request.form.get('department')
        is_active = request.form.get('is_active') == 'on'

        if employee:
            # Editing existing employee
            # Check if username changed and if new username exists
            if employee.username != username:
                if User.query.filter_by(username=username).first():
                    return render_template('add_employee.html', employee=employee, error='Username already exists')
            
            # Check if email changed and if new email exists
            if employee.email != email:
                if User.query.filter_by(email=email).first():
                    return render_template('add_employee.html', employee=employee, error='Email already exists')
            
            employee.username = username
            employee.email = email
            employee.full_name = full_name
            employee.department = department
            employee.is_active = is_active
            
            # Only update password if provided
            if password:
                employee.set_password(password)
                db.session.commit()
                # Send password change notification
                send_password_change_email(employee)
            else:
                db.session.commit()
        else:
            # Adding new employee
            if User.query.filter_by(username=username).first():
                return render_template('add_employee.html', error='Username already exists')
            
            if User.query.filter_by(email=email).first():
                return render_template('add_employee.html', error='Email already exists')

            employee = User(
                username=username,
                email=email,
                full_name=full_name,
                department=department,
                role='employee',
                is_active=True
            )
            employee.set_password(password)
            db.session.add(employee)
            db.session.commit()
            
            # Send welcome email with credentials
            send_welcome_email(employee, password)
        
        db.session.commit()
        return redirect(url_for('manage_employees'))

    return render_template('add_employee.html', employee=employee)


@app.route('/admin/rotas')
@login_required
def manage_rotas():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    employees = User.query.filter_by(role='employee').all()
    return render_template('manage_rotas.html', employees=employees)


@app.route('/admin/employee/<int:employee_id>/rotas', methods=['GET', 'POST'])
@login_required
def employee_rotas(employee_id):
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    employee = User.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        day_of_week = request.form.get('day_of_week')
        shift_start = request.form.get('shift_start')
        shift_end = request.form.get('shift_end')
        
        # Check if rota already exists for this day
        existing_rota = Rota.query.filter_by(
            user_id=employee_id,
            day_of_week=day_of_week,
            is_active=True
        ).first()
        
        if existing_rota:
            # Update existing rota
            existing_rota.shift_start = datetime.strptime(shift_start, '%H:%M').time()
            existing_rota.shift_end = datetime.strptime(shift_end, '%H:%M').time()
        else:
            # Create new rota
            rota = Rota(
                user_id=employee_id,
                day_of_week=day_of_week,
                shift_start=datetime.strptime(shift_start, '%H:%M').time(),
                shift_end=datetime.strptime(shift_end, '%H:%M').time()
            )
            db.session.add(rota)
        
        db.session.commit()
        return redirect(url_for('employee_rotas', employee_id=employee_id))
    
    rotas = Rota.query.filter_by(user_id=employee_id, is_active=True).order_by(
        db.case(
            (Rota.day_of_week == 'Monday', 1),
            (Rota.day_of_week == 'Tuesday', 2),
            (Rota.day_of_week == 'Wednesday', 3),
            (Rota.day_of_week == 'Thursday', 4),
            (Rota.day_of_week == 'Friday', 5),
            (Rota.day_of_week == 'Saturday', 6),
            (Rota.day_of_week == 'Sunday', 7),
        )
    ).all()
    
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    return render_template('employee_rotas.html', employee=employee, rotas=rotas, days_of_week=days_of_week)


@app.route('/admin/rota/<int:rota_id>/delete', methods=['POST'])
@login_required
def delete_rota(rota_id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    rota = Rota.query.get_or_404(rota_id)
    employee_id = rota.user_id
    db.session.delete(rota)
    db.session.commit()
    
    return redirect(url_for('employee_rotas', employee_id=employee_id))


@app.route('/admin/attendance-records')
@login_required
def attendance_records():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    page = request.args.get('page', 1, type=int)
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)
    
    query = Attendance.query
    
    if date_from:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= from_date)
    
    if date_to:
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= to_date)
    
    records = query.order_by(Attendance.date.desc(), Attendance.check_in.desc()).paginate(page=page, per_page=15)
    
    return render_template('attendance_records.html', records=records, date_from=date_from, date_to=date_to)


@app.route('/admin/reports')
@login_required
def reports():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    report_type = request.args.get('type', 'monthly', type=str)
    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)
    
    report_data = {}
    
    if report_type == 'monthly':
        report_data = get_monthly_report(month, year)
    elif report_type == 'employee':
        report_data = get_employee_summary_report()
    elif report_type == 'working_hours':
        report_data = get_working_hours_report(month, year)
    elif report_type == 'absence':
        report_data = get_absence_report(month, year)
    
    return render_template('reports.html', 
                         report_type=report_type,
                         report_data=report_data,
                         month=month,
                         year=year)


@app.route('/api/admin/stats')
@login_required
def get_stats():
    if current_user.role != 'admin':
        return jsonify({'success': False}), 403
    
    today = datetime.utcnow().date()
    
    stats = {
        'total_employees': User.query.filter_by(role='employee').count(),
        'present_today': db.session.query(Attendance).filter(
            Attendance.date == today,
            Attendance.status == 'present'
        ).count(),
        'absent_today': db.session.query(Attendance).filter(
            Attendance.date == today,
            Attendance.status == 'absent'
        ).count(),
    }
    
    return jsonify(stats)


@app.route('/api/admin/employee-hours-today')
@login_required
def get_employee_hours_today():
    if current_user.role != 'admin':
        return jsonify({'success': False}), 403
    
    today = datetime.utcnow().date()
    
    # Get all employees
    employees = User.query.filter_by(role='employee', is_active=True).all()
    
    employee_hours = []
    for emp in employees:
        attendance = Attendance.query.filter_by(
            user_id=emp.id,
            date=today
        ).first()
        
        hours_worked = "0h 0m"
        status = "Not Checked In"
        
        if attendance:
            if attendance.check_in and attendance.check_out:
                time_diff = attendance.check_out - attendance.check_in
                hours = int(time_diff.total_seconds() // 3600)
                minutes = int((time_diff.total_seconds() % 3600) // 60)
                hours_worked = f"{hours}h {minutes}m"
                status = "Checked Out"
            elif attendance.check_in:
                # Calculate current hours if still checked in
                now = datetime.utcnow()
                time_diff = now - attendance.check_in
                hours = int(time_diff.total_seconds() // 3600)
                minutes = int((time_diff.total_seconds() % 3600) // 60)
                hours_worked = f"{hours}h {minutes}m"
                status = "Working"
        
        employee_hours.append({
            'name': emp.full_name,
            'hours': hours_worked,
            'status': status
        })
    
    # Sort by hours worked (descending)
    employee_hours.sort(key=lambda x: x['hours'], reverse=True)
    
    return jsonify({'employees': employee_hours})


# ===================== Excel Export Routes =====================
@app.route('/admin/export/monthly-report')
@login_required
def export_monthly_report():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)
    
    report_data = get_monthly_report(month, year)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"
    
    # Header
    ws['A1'] = f"Monthly Attendance Report - {['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month-1]} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Summary Section
    ws['A3'] = "Summary"
    ws['A3'].font = Font(bold=True, size=12)
    
    summary_data = [
        ['Metric', 'Count'],
        ['Total Present', report_data['summary']['total_present']],
        ['Total Absent', report_data['summary']['total_absent']],
        ['Total Leave', report_data['summary']['total_leave']],
        ['Total Records', report_data['summary']['total_records']],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Daily Breakdown
    ws['A10'] = "Daily Breakdown"
    ws['A10'].font = Font(bold=True, size=12)
    
    daily_headers = ['Date', 'Day', 'Present', 'Absent', 'Leave']
    for col_idx, header in enumerate(daily_headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row_idx, (date, stats) in enumerate(report_data['daily_stats'].items(), start=12):
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        day_name = date_obj.strftime('%A')
        ws.cell(row=row_idx, column=1, value=date)
        ws.cell(row=row_idx, column=2, value=day_name)
        ws.cell(row=row_idx, column=3, value=stats['present'])
        ws.cell(row=row_idx, column=4, value=stats['absent'])
        ws.cell(row=row_idx, column=5, value=stats['leave'])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name=f'Monthly_Report_{month}_{year}.xlsx')


@app.route('/admin/export/employee-report')
@login_required
def export_employee_report():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    report_data = get_employee_summary_report()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Summary"
    
    # Header
    ws['A1'] = "Employee Attendance Summary Report (All Time)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    headers = ['Employee Name', 'Username', 'Department', 'Present', 'Absent', 'Leave', 'Total Records', 'Total Hours']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row_idx, emp in enumerate(report_data['employees'], start=4):
        ws.cell(row=row_idx, column=1, value=emp['name'])
        ws.cell(row=row_idx, column=2, value=emp['username'])
        ws.cell(row=row_idx, column=3, value=emp['department'])
        ws.cell(row=row_idx, column=4, value=emp['total_present'])
        ws.cell(row=row_idx, column=5, value=emp['total_absent'])
        ws.cell(row=row_idx, column=6, value=emp['total_leave'])
        ws.cell(row=row_idx, column=7, value=emp['total_records'])
        ws.cell(row=row_idx, column=8, value=emp['total_hours'])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Employee_Summary_Report.xlsx')


@app.route('/admin/export/working-hours-report')
@login_required
def export_working_hours_report():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)
    
    report_data = get_working_hours_report(month, year)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Working Hours"
    
    # Header
    month_name = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month-1]
    ws['A1'] = f"Working Hours Report - {month_name} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Summary
    ws['A3'] = "Summary"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = "Total Hours:"
    ws['B4'] = report_data['total_hours']
    ws['A5'] = "Average Hours:"
    ws['B5'] = report_data['average_hours']
    
    # Employee Data
    ws['A7'] = "Employee Working Hours"
    ws['A7'].font = Font(bold=True, size=12)
    
    headers = ['Employee Name', 'Username', 'Department', 'Working Days', 'Total Hours', 'Average Hours/Day']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row_idx, emp in enumerate(report_data['employees'], start=9):
        ws.cell(row=row_idx, column=1, value=emp['name'])
        ws.cell(row=row_idx, column=2, value=emp['username'])
        ws.cell(row=row_idx, column=3, value=emp['department'])
        ws.cell(row=row_idx, column=4, value=emp['working_days'])
        ws.cell(row=row_idx, column=5, value=emp['total_hours'])
        ws.cell(row=row_idx, column=6, value=emp['average_hours'])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Working_Hours_Report_{month}_{year}.xlsx')


@app.route('/admin/export/absence-report')
@login_required
def export_absence_report():
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    
    month = request.args.get('month', datetime.utcnow().month, type=int)
    year = request.args.get('year', datetime.utcnow().year, type=int)
    
    report_data = get_absence_report(month, year)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Absence Report"
    
    # Header
    month_name = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month-1]
    ws['A1'] = f"Absence Report - {month_name} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Summary
    ws['A3'] = f"Total Absences: {report_data['total_absences']}"
    ws['A3'].font = Font(bold=True)
    
    # Absence Data
    headers = ['Employee Name', 'Username', 'Department', 'Date', 'Day', 'Notes']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for row_idx, absence in enumerate(report_data['absences'], start=6):
        ws.cell(row=row_idx, column=1, value=absence['employee_name'])
        ws.cell(row=row_idx, column=2, value=absence['username'])
        ws.cell(row=row_idx, column=3, value=absence['department'])
        ws.cell(row=row_idx, column=4, value=absence['date'])
        ws.cell(row=row_idx, column=5, value=absence['day'])
        ws.cell(row=row_idx, column=6, value=absence['notes'])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'Absence_Report_{month}_{year}.xlsx')


# ===================== Report Generation Functions =====================
def get_monthly_report(month, year):
    """Generate monthly attendance summary report"""
    from calendar import monthrange
    
    first_day = datetime(year, month, 1).date()
    last_day = datetime(year, month, monthrange(year, month)[1]).date()
    
    records = Attendance.query.filter(
        Attendance.date >= first_day,
        Attendance.date <= last_day
    ).all()
    
    summary = {
        'total_records': len(records),
        'total_present': len([r for r in records if r.status == 'present']),
        'total_absent': len([r for r in records if r.status == 'absent']),
        'total_leave': len([r for r in records if r.status == 'leave']),
    }
    
    # Daily breakdown
    daily_stats = {}
    for day in range(1, monthrange(year, month)[1] + 1):
        date = datetime(year, month, day).date()
        day_records = [r for r in records if r.date == date]
        daily_stats[date.strftime('%Y-%m-%d')] = {
            'present': len([r for r in day_records if r.status == 'present']),
            'absent': len([r for r in day_records if r.status == 'absent']),
            'leave': len([r for r in day_records if r.status == 'leave']),
        }
    
    return {
        'summary': summary,
        'daily_stats': daily_stats,
        'month': month,
        'year': year
    }


def get_employee_summary_report():
    """Generate employee-wise attendance summary"""
    employees = User.query.filter_by(role='employee').all()
    
    employee_stats = []
    for emp in employees:
        records = Attendance.query.filter_by(user_id=emp.id).all()
        
        total_present = len([r for r in records if r.status == 'present'])
        total_absent = len([r for r in records if r.status == 'absent'])
        total_leave = len([r for r in records if r.status == 'leave'])
        
        # Calculate total hours
        total_hours = 0
        for record in records:
            if record.check_in and record.check_out:
                duration = record.check_out - record.check_in
                total_hours += duration.total_seconds() / 3600
        
        employee_stats.append({
            'id': emp.id,
            'name': emp.full_name,
            'username': emp.username,
            'department': emp.department or '-',
            'total_present': total_present,
            'total_absent': total_absent,
            'total_leave': total_leave,
            'total_records': len(records),
            'total_hours': round(total_hours, 2)
        })
    
    return {'employees': employee_stats}


def get_working_hours_report(month, year):
    """Generate working hours report for the month"""
    from calendar import monthrange
    
    first_day = datetime(year, month, 1).date()
    last_day = datetime(year, month, monthrange(year, month)[1]).date()
    
    employees = User.query.filter_by(role='employee').all()
    
    working_hours_data = []
    for emp in employees:
        records = Attendance.query.filter_by(user_id=emp.id).filter(
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        
        total_hours = 0
        working_days = 0
        for record in records:
            if record.check_in and record.check_out:
                duration = record.check_out - record.check_in
                total_hours += duration.total_seconds() / 3600
                working_days += 1
        
        if working_days > 0:
            avg_hours = total_hours / working_days
        else:
            avg_hours = 0
        
        working_hours_data.append({
            'name': emp.full_name,
            'username': emp.username,
            'department': emp.department or '-',
            'working_days': working_days,
            'total_hours': round(total_hours, 2),
            'average_hours': round(avg_hours, 2)
        })
    
    total_all_hours = sum([e['total_hours'] for e in working_hours_data])
    avg_all_hours = sum([e['average_hours'] for e in working_hours_data]) / len(employees) if employees else 0
    
    return {
        'employees': working_hours_data,
        'total_hours': round(total_all_hours, 2),
        'average_hours': round(avg_all_hours, 2),
        'month': month,
        'year': year
    }


def get_absence_report(month, year):
    """Generate absence report for the month"""
    from calendar import monthrange
    
    first_day = datetime(year, month, 1).date()
    last_day = datetime(year, month, monthrange(year, month)[1]).date()
    
    absences = Attendance.query.filter(
        Attendance.date >= first_day,
        Attendance.date <= last_day,
        Attendance.status == 'absent'
    ).all()
    
    absence_data = []
    for absence in absences:
        employee = User.query.get(absence.user_id)
        absence_data.append({
            'employee_name': employee.full_name,
            'username': employee.username,
            'department': employee.department or '-',
            'date': absence.date.strftime('%Y-%m-%d'),
            'day': absence.date.strftime('%A'),
            'notes': absence.notes or '-'
        })
    
    return {
        'absences': absence_data,
        'total_absences': len(absence_data),
        'month': month,
        'year': year
}

# ===================== Database Utilities =====================
def create_default_admin():
    """Create default admin user if it doesn't exist"""
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            email='admin@attendance.com',
            full_name='Administrator',
            role='admin',
            department='HR'
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()


@app.cli.command('flush-db')
@click.option('--force', is_flag=True, help='Do not prompt for confirmation.')
@click.option('--keep-admin/--no-keep-admin', default=True, help='Recreate default admin after flush.')
@with_appcontext
def flush_db_command(force: bool, keep_admin: bool):
    """Permanently delete ALL data and recreate tables.
    By default, recreates the default admin account (admin/admin123).
    """
    if not force:
        if not click.confirm('This will permanently delete ALL data. Continue?'):
            click.echo('Aborted.')
            return

    # Drop and recreate all tables
    db.drop_all()
    db.create_all()

    if keep_admin:
        create_default_admin()

    click.echo('Database flushed successfully.')
    click.echo(f"Users: {User.query.count()}, Attendance: {Attendance.query.count()}, Rotas: {Rota.query.count()}")
def init_db():
    with app.app_context():
        db.create_all()
        create_default_admin()
        print("Ensured default admin exists (admin/admin123)")


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
